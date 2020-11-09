import discord
from discord.ext import commands
from discord.utils import get
import random as rnd
import openpyxl
import asyncio
import datetime as dt
import math

#token.txt is a file not uploaded to git, containing the bot token
with open("token.txt", 'r') as f:
    token = f.readline()[:-1]
    botID = f.readline()

client = discord.Client()
bot = commands.Bot(command_prefix='~', case_insensitive = True)
bot.remove_command('help')

'''
player.currentVote [x,y]
x:
0 : Non-vote
1 : Yes
2 : No
-1 : Forfeit
-2 : Inactive player

y: Order of votes

Turn is the index of the current players turn, or the next player in the case where the game is between turns
It skips players who have left the game, and loops back to 0
globalTurn is the actual turn of the game. It only increments by 1

state
0 : The previous turn has ended, and the bot is waiting for historians to formalise the end of the turn and start the next
1 : The current player is writing a proposal to to be discussed and voted on
2 : A proposal has been made and player can vote for it


turn.end
0 : All votes
1 : Sufficient votes
2 : Out of voting time
3 : Out of proposal time
4 : Passed
'''


players = []
class Player(object):
    def __init__(self, discObj, globalTurn):
        self.discord = discObj
        self.name = None
        self.points = 0
        self.botcoins = 0
        self.active = True
        self.lastMessage = None
        self.currentVote = Vote(None, None, None, self)
        self.online = True
        statNames = ['messages','daysPlaying','daysOnline','proposals','firstVotes','lastVotes','bruh','roulette']
        self.stats = {i : 0 for i in statNames}
    def __repr__(self):
        return self.name

class Vote(object):
    def __init__(self, value, order, time, player):
        self.value = value
        self.order = order
        self.time = time
        self.player = player
    def __repr__(self):
        if self.value == 0: return self.player.name + ' No Vote'
        elif self.value == 1: return self.player.name + ' #' + str(self.order) + ': Yes'
        elif self.value == 2: return self.player.name + ' #' + str(self.order) + ': No'
        elif self.value == -2: return self.player.name + ' Inactive'
        elif self.value == None: return self.player.name + ' None'


turns = []
class Turn(object):
    def __init__(self, turn):
        self.turnNumber = turn
        self.proponent = None
        self.passed = None
        self.end = None
        self.voteHistory = [Vote(None, None, None, None)] * len(players)
    def __repr__(self):
        return str(self.turnNumber) + ': ' + str(self.proponent.name)


game = None
class Parameters(object):
    def __init__(self):
        self.turn = None
        self.globalTurn = None
        self.state = None
        self.proposalTime = None
        self.votingTime = None
        self.yesProportion = [None,None]
        self.timeoutProportion = [None,None]
        self.timeoutMinimum = [None,None]
        self.timerEnd = None
        self.firstVote = None
        self.lastVote = None
        self.ready = None
        self.rolled = None
        self.poolNumber = None
        self.botcoins = None
    def __repr__(self):
        return 'Turn:' + str(self.globalTurn) + '  State:' + str(self.state)


rulePool = []
ruleQueue = []
class Rule(object):
    def __init__(self, number):
        self.number = number
        self.text = None
        self.proposer = None
        self.active = True
        self.turn = None
        self.msgID = None
        self.down = 0
        self.up = 0
    def __repr__(self):
        return str(self.number) + ': ' + self.proposer.name


minigame = None
class Minigame(object):
    def __init__(self):
        self.rouletteState = None
        self.rouletteTimer = None
        self.rouletteNumber = 0
        self.rouletteChamber = 0


#Excel sheets
wb = openpyxl.load_workbook('nomic.xlsx')
ws1 = wb['Players']
ws2 = wb['Turns']
ws3 = wb['Misc']
ws4 = wb['Rules']
ws5 = wb['Pool']

bc = openpyxl.load_workbook('botcoins.xlsx')
bc1 = bc['Sheet1']

def loadData():
    global game, players, turns, ruleQueue, rulePool, minigame, summaryMsg, rouletteMsg
    if game is not None:
        return
    game = Parameters()
    game.turn = ws3['B3'].value
    game.globalTurn = ws3['B4'].value
    game.state = ws3['B5'].value
    game.firstVote = ws3['B7'].value
    try:
        game.lastVote = get(players, discord__id = int(ws3['B8'].value))
    except TypeError:
        game.lastVote = None
    game.voteNumber = ws3['B9'].value
    game.proposalTime = ws3['B11'].value
    game.votingTime = ws3['B12'].value
    game.yesProportion = [float(i) for i in ws3['B13'].value.split(',')]
    game.timeoutProportion = [float(i) for i in ws3['B14'].value.split(',')]
    game.timeoutMinimum = [float(i) for i in ws3['B15'].value.split(',')]
    game.transmute = ws3['B16'].value
    game.timerEnd = ws3['B17'].value
    game.ready = ws3['B18'].value
    summaryMsg = ws3['B20'].value
    game.rolled = ws3['B21'].value
    game.poolNumber = ws5['P2'].value
    game.botcoins = bc1['A3'].value

    for i in range(ws3['B1'].value):
        nextPlayer = get(nomicServer.members, id=int(ws1.cell(3, i+2).value))
        nextPlayer = Player(nextPlayer, game.globalTurn)
        if nextPlayer.discord is None:
            nextPlayer.discord = ws1.cell(1, i+2).value
        nextPlayer.name = ws1.cell(1, i+2).value
        nextPlayer.active = ws1.cell(5, i+2).value
        nextPlayer.lastMessage = ws1.cell(6, i+2).value
        nextPlayer.points = ws1.cell(7, i+2).value
        nextPlayer.botcoins = bc1.cell(3, i+2).value
        vote = ws1.cell(8,i+2).value
        if vote is None:
            nextPlayer.currentVote = Vote(None, '', '', nextPlayer)
        else:
            vote = vote.split(',')
            vote[0] = int(vote[0])
            if vote[1] != '':
                vote[1] = int(vote[1])
            if vote[2] != '':
                vote[2] = dt.datetime.strptime(vote[2], '%d/%m/%Y %H:%M:%S')
            nextPlayer.currentVote = Vote(vote[0], vote[1], vote[2], nextPlayer)
        nextPlayer.online = ws1.cell(9, i+2).value
        j = 12
        for k in nextPlayer.stats:
            nextPlayer.stats[k] = ws1.cell(j, i+2).value
            j += 1
        players.append(nextPlayer)

    for i in range(game.globalTurn):
        nextTurn = Turn(i)
        nextTurn.proponent = get(players, discord__id=int(ws2.cell(i+3, 2).value))
        if nextTurn.proponent is None:
            nextTurn.proponent = ws2.cell(i+3, 3).value
        nextTurn.passed = ws2.cell(i+3, 4).value
        nextTurn.end = ws2.cell(i+3, 5).value
        for j in range(len(players)):
            vote = ws2.cell(i+3,j+6).value
            if vote is None:
                nextTurn.voteHistory[j] = Vote(None, '', '', players[j])
            else:
                vote = vote.split(',')
                vote[0] = int(vote[0])
                if vote[1] != '':
                    vote[1] = int(vote[1])
                if vote[2] != '':
                    vote[2] = dt.datetime.strptime(vote[2], '%d/%m/%Y %H:%M:%S')
                nextTurn.voteHistory[j] = Vote(vote[0], vote[1], vote[2], players[j])
        turns.append(nextTurn)

    i = 2
    while ws5.cell(i,1).value is not None:
        ruleQueue.append(Rule(ws5.cell(i,1).value))
        ruleQueue[-1].text = ws5.cell(i,2).value
        ruleQueue[-1].proposer = get(players, discord__id=int(ws5.cell(i,3).value))
        ruleQueue[-1].msgID = int(ws5.cell(i,4).value)
        ruleQueue[-1].down = ws5.cell(i,5).value
        i += 1

    i = 2
    while ws5.cell(i,7).value is not None:
        rulePool.append(Rule(ws5.cell(i,7).value))
        rulePool[-1].text = ws5.cell(i,8).value
        rulePool[-1].proposer = get(players, discord__id=int(ws5.cell(i,9).value))
        rulePool[-1].active = ws5.cell(i,10).value
        rulePool[-1].turn = ws5.cell(i,11).value
        rulePool[-1].msgID = int(ws5.cell(i,12).value)
        rulePool[-1].up = ws5.cell(i,13).value
        rulePool[-1].down = ws5.cell(i,14).value
        i += 1

    minigame = Minigame()
    minigame.rouletteState = ws3['B23'].value
    minigame.rouletteTimer = ws3['B24'].value
    minigame.rouletteNumber = sum([player.stats['roulette']>-1 for player in players])
    minigame.rouletteChamber = ws3['B25'].value
    rouletteMsg = ws3['B26'].value
    if minigame.rouletteState == 0:
        asyncio.create_task(rouletteWait())
    elif minigame.rouletteState < 3:
        asyncio.create_task(rouletteSignup())
    else:
        global rouletteTask
        rouletteTask = asyncio.create_task(roulettePlay())


    if game.state == 1:
        global proposalTask
        proposalTask = asyncio.create_task(proposalTimeLimit(game.timerEnd))
    elif game.state == 2:
        global voteTask
        voteTask = asyncio.create_task(votingTimeLimit(game.timerEnd))

async def saveData():
    ws3['B1'] = len(players)
    ws3['B3'] = game.turn
    ws3['B4'] = game.globalTurn
    ws3['B5'] = game.state
    ws3['B9'] = game.voteNumber
    ws3['B16'] = game.transmute
    ws3['B17'] = game.timerEnd
    ws3['B18'] = game.ready
    if summaryMsg: ws3['B20'] = str(summaryMsg.id)
    else: ws3['B20'] = None
    ws3['B21'] = game.rolled
    ws5['P2'] = game.poolNumber
    bc1['A3'] = game.botcoins
    ws3['B23'] = minigame.rouletteState
    ws3['B24'] = minigame.rouletteTimer
    ws3['B25'] = minigame.rouletteChamber
    if rouletteMsg: ws3['B26'] = str(rouletteMsg.id)
    else: ws3['B26'] = None
    for player in players:
        i = players.index(player)
        if player.discord in nomicServer.members:
            ws1.cell(2, i+2, player.discord.name)
            ws1.cell(3, i+2, str(player.discord.id))
        ws1.cell(5, i+2, player.active)
        ws1.cell(6, i+2, player.lastMessage)
        ws1.cell(7, i+2, player.points)
        bc1.cell(3, i+2, player.botcoins)
        if player.currentVote.value is None:
            ws1.cell(8, i+2).value = None
        elif player.currentVote.time == '':
            ws1.cell(8, i+2, str(player.currentVote.value) + ',' + str(player.currentVote.order) + ',')
        else:
            ws1.cell(8, i+2, str(player.currentVote.value) + ',' + str(player.currentVote.order) + ',' + dt.datetime.strftime(player.currentVote.time, '%d/%m/%Y %H:%M:%S'))
        ws1.cell(9, i+2, player.online)
    ws3['B7'] = game.firstVote
    if game.lastVote is not None:
        ws3['B8'] = str(game.lastVote.discord.id)
    else: ws3['B8'] = None

    for i in range(game.globalTurn):
        j = 6
        for vote in turns[i].voteHistory:
            if vote.value is None:
                ws2.cell(i+3, j).value = None
            elif vote.time == '':
                ws2.cell(i+3, j, str(vote.value) + ',' + str(vote.order) + ',')
            else:
                ws2.cell(i+3, j, str(vote.value) + ',' + str(vote.order) + ',' + dt.datetime.strftime(vote.time, '%d/%m/%Y %H:%M:%S'))
            j += 1
        ws2.cell(i+3, 1, turns[i].turnNumber)
        proponent = turns[i].proponent
        if not isinstance(proponent, str):
            ws2.cell(i+3, 2, str(proponent.discord.id))
            ws2.cell(i+3, 3, proponent.name)
        ws2.cell(i+3, 4, turns[i].passed)
        ws2.cell(i+3, 5, turns[i].end)
    for player in players:
        i = 12
        for stat in player.stats.values():
            ws1.cell(i, players.index(player)+2, stat)
            i += 1

    for i in range(len(ruleQueue)):
        rule = ruleQueue[i]
        ws5.cell(i+2, 1).value = rule.number
        ws5.cell(i+2, 2).value = rule.text
        ws5.cell(i+2, 3).value = str(rule.proposer.discord.id)
        ws5.cell(i+2, 4).value = str(rule.msgID)
        ws5.cell(i+2, 5).value = rule.down

    for i in range(len(rulePool)):
        rule = rulePool[i]
        ws5.cell(i+2, 7, rule.number)
        ws5.cell(i+2, 8, rule.text)
        ws5.cell(i+2, 9, str(rule.proposer.discord.id))
        ws5.cell(i+2, 10, rule.active)
        ws5.cell(i+2, 11, rule.turn)
        ws5.cell(i+2, 12, str(rule.msgID))
        ws5.cell(i+2, 13, rule.up)
        ws5.cell(i+2, 14, rule.down)

    wb.save('nomic.xlsx')
    bc.save('botcoins.xlsx')
    print("Saved " + str(dt.datetime.now()))

setup = False
@bot.event
async def on_ready():
    global setup
    if setup: return
    setup = True
    global nomicServer, botMember, botChannel, histBotChannel, updateChannel, votingChannel, playerRole
    #Commonly used channels and roles
    nomicServer = get(bot.guilds, name='Nomic')
    botMember = get(nomicServer.members, id=int(botID))

    botChannel = get(nomicServer.channels, name='bot-commands')
    histBotChannel = get(nomicServer.channels, name='historian-bot')
    updateChannel = get(nomicServer.channels, name='game-updates')
    votingChannel = get(nomicServer.channels, name='voting')

    playerRole = get(nomicServer.roles, name='Player')

    loadData()
    global summaryMsg, rouletteMsg
    if summaryMsg:
        summaryMsg = await votingChannel.fetch_message(int(summaryMsg))
    if rouletteMsg:
        rouletteMsg = await get(nomicServer.channels,name='the-roulette-games').fetch_message(int(rouletteMsg))

    #Initial state role
    await bot.change_presence(activity=discord.Game(name='~help'))
    roleNames = ['Game State: Waiting', 'Game State: Proposing', 'Game State: Voting']
    await botMember.add_roles(get(nomicServer.roles, name=roleNames[game.state]))

    
    #if game.state == 2:
    #    await checkVotes(0)
    print("Bot is ready")



@bot.command()
async def save(ctx):
    if not ctx.channel == histBotChannel: return
    await saveData()

@bot.command()
async def data(ctx):
    if not ctx.channel == botChannel: return
    await saveData()    
    await botChannel.send(file=discord.File('nomic.xlsx'))



@bot.command()
async def join(ctx):
    global players
    if ctx.channel == get(nomicServer.channels,name='the-roulette-games'):
        await joinRoulette(ctx)
        return
    elif ctx.channel != botChannel: return
    player = get(players, discord=ctx.author)
    if player is not None:
        #When the author is/was part of the game
        index = players.index(get(players, discord=ctx.author))
        if player.active:
            await botChannel.send("You are already a player")
        return
    #New Player
    await botChannel.send("{} has joined the game!".format(ctx.author.mention))
    await ctx.author.add_roles(playerRole)
    newPlayerObj = Player(ctx.author, game.globalTurn)
    newPlayerObj.lastMessage = dt.datetime.now()
    if len(players) == 0:
        newPlayer(0, newPlayerObj)
        await saveData()
        return
    #Pick a random place to insert the player into
    placement = rnd.randint(0,len(players)-1)
    if placement == 0 and rnd.randint(0,1) == 0:
        #The first and last position are equivalent, so there is a 50% chance of each
        placement = len(players)
    newPlayer(placement, newPlayerObj)
    #Find the players before and after who are still part of the game
    i = placement
    while not players[i-1].active:
        i -= 1
        if i == -1:
            i = len(players) -1
    before = players[i-1].name
    i = placement
    if i == len(players)-1:
        i = -1
    while not players[i+1].active:
        i += 1
        if i == len(players)-1:
            i = -1
    after = players[i+1].name
    if len(players) > 2:
        await botChannel.send("You are player #{} in the turn order, between {} & {}".format(placement+1, before, after))
    await saveData()

def newPlayer(index, player):
    global players, game
    #If the new player is before the current player, increment turn unless the game hasn't started yet
    if index <= game.turn and not (game.globalTurn == 1 and game.state == 0):
        game.turn += 1
    if index == len(players):
        players = players + [player]
    else:
        players = players[:index] + [player] + players[index:]



@bot.command()
async def ready(ctx):
    global game
    if not (ctx.channel == botChannel and game.state == 0 and get(nomicServer.roles,name='Historian') in ctx.author.roles): return
    game.ready = True
    await start()
    await saveData()

@bot.command()
async def pause(ctx):
    if not (ctx.channel == botChannel and get(nomicServer.roles,name='Historian') in ctx.author.roles):
        return
    if game.state == 1:
        global proposalTask
        proposalTask.cancel()
        game.state == 0
        await botMember.add_roles(get(nomicServer.roles, name='Game State: Waiting'))
        await botMember.remove_roles(get(nomicServer.roles, name='Game State: Proposing'))
        await players[game.turn].discord.add_roles(get(nomicServer.roles, name='Next Player'))
        await players[game.turn].discord.remove_roles(get(nomicServer.roles, name='Current Player'))
        nextTurn = game.turn + 1
        if nextTurn >= len(players):
            nextTurn = 0
        while not players[nextTurn].active:
            nextTurn += 1
            if nextTurn > len(players):
                nextTurn = 0
        await players[nextTurn].discord.remove_roles(get(nomicServer.roles, name='Next Player'))
    elif game.state == 2:
        game.ready = False

async def start():
    global game
    #Begin the proposal phase
    game.state = 1
    #Give roles
    await updateChannel.send("Turn #{}! {}'s turn has begun, make a proposal using ~propose".format(game.globalTurn+1, players[game.turn].discord.mention))
    #await players[game.turn].discord.add_roles(get(nomicServer.roles, name='Current Player'))
    #await players[game.turn].discord.remove_roles(get(nomicServer.roles, name='Next Player'))
    nextTurn = game.turn + 1
    if nextTurn >= len(players):
        nextTurn = 0
    while not players[nextTurn].active:
        nextTurn += 1
        if nextTurn > len(players):
            nextTurn = 0
    #await players[nextTurn].discord.add_roles(get(nomicServer.roles, name='Next Player'))
    #await botMember.add_roles(get(nomicServer.roles, name='Game State: Proposing'))
    #await botMember.remove_roles(get(nomicServer.roles, name='Game State: Waiting'))
    #Begin timer for proposing
    game.timerEnd = dt.datetime.now() + dt.timedelta(seconds = game.proposalTime)
    global proposalTask
    proposalTask = asyncio.create_task(proposalTimeLimit(game.timerEnd))

    global ruleQueue, rulePool
    queueChannel = get(nomicServer.channels, name='rule-pool-queue')
    poolChannel = get(nomicServer.channels, name='rule-pool')
    i = 2
    for rule in ruleQueue:
        msg = await queueChannel.fetch_message(rule.msgID)
        await msg.delete()
        game.poolNumber += 1
        rulePool.append(Rule(game.poolNumber))
        rulePool[-1].text = rule.text
        rulePool[-1].proposer = rule.proposer
        newMsg = await poolChannel.send('#{}: {}\n{}'.format(game.poolNumber,rule.proposer,rule.text))
        rulePool[-1].msgID = newMsg.id
        await newMsg.add_reaction('\U0001F6D1')
        await newMsg.add_reaction('\U0001F53C')
        ws5.cell(i, 1).value = None
        ws5.cell(i, 2).value = None
        ws5.cell(i, 3).value = None
        ws5.cell(i, 4).value = None
        ws5.cell(i, 5).value = None
    ruleQueue = []
    wb.save('nomic.xlsx')



async def proposalTimeLimit(end):
    now = dt.datetime.now()
    if (end - now).total_seconds() > 3601:
        await asyncio.sleep((end - now).total_seconds() -3600)
        await votingChannel.send("{}, you have one hour left to propose".format(players[game.turn].discord.mention))
        await asyncio.sleep(3600)
        await updateChannel.send("A proposal was not made in time, waiting for the next turn")
    else:
        await asyncio.sleep((end - now).total_seconds())
        await updateChannel.send("A proposal was not made in time, waiting for the next turn")
    await endTurn(0, 3)

async def votingTimeLimit(end):
    global game
    now = dt.datetime.now()
    if (end - now).total_seconds() > 3601:
        await asyncio.sleep((end - now).total_seconds() -3600)
        toVoteRole = get(nomicServer.roles, name='To Vote')
        await votingChannel.send("{}, you have one hour left to vote".format(toVoteRole.mention))
        await asyncio.sleep(3600)
        await votingChannel.send("Voting time is up")
    else:
        await asyncio.sleep((end - now).total_seconds())
        await votingChannel.send("Voting time is up")
    game.lastVote = None
    await checkVotes(1)

@bot.command()
async def timeout(ctx):
    if not(get(nomicServer.roles, name='Historian') in ctx.author.roles and ctx.channel == botChannel): return
    if game.state == 1:
        await updateChannel.send("A proposal was not made in time, waiting for the next turn")
        global proposalTask
        proposalTask.cancel()
        await endTurn(0, 3)
    elif game.state == 2:
        await votingChannel.send("Voting time is up")
        game.lastVote = None
        global voteTask
        voteTask.cancel()
        await checkVotes(1)


@bot.command(name='pass')
async def passTurn(ctx):
    global game
    if not(ctx.author == players[game.turn].discord and ctx.channel == botChannel and game.state == 1): return
    await updateChannel.send('The current turn has been passed, waiting for the next turn to start')
    global proposalTask
    proposalTask.cancel()
    if game.rolled:
        global rollTask
        rollTask.cancel()
        game.rolled = False
    await endTurn(0,4)



@bot.command()
async def propose(ctx):
    global players, game
    game.transmute = 0
    if not (ctx.channel == votingChannel and game.state == 1 and (ctx.author == players[game.turn].discord or ctx.author == botMember)): return
    game.state = 2
    #FirstVote is whether or not the first vote has been made yet, lastVote is the index of the most recent vote
    game.firstVote = False
    game.lastVote = None
    toVoteRole = get(nomicServer.roles, name='To Vote')
    for player in players:
        if player.active:
            player.currentVote = Vote(0, '', '', player)
            await player.discord.add_roles(toVoteRole)
        else:
            player.currentVote = Vote(-2, '', '', player)
    game.voteNumber = 0
    players[game.turn].stats['proposals'] += 1
    #End timer
    global proposalTask
    proposalTask.cancel()
    if game.rolled:
        global rollTask, rulePool
        rollTask.cancel()
        rule = rulePool[game.rolled-1]
        game.rolled = None
        rule.active = False
        rule.turn = game.globalTurn
        await get(nomicServer.channels, name='rule-pool').fetch_message(rule.msgID).delete()
    #Begin voting phase
    instPass = math.ceil(sum([x.active for x in players])*game.yesProportion[0])
    instFail = math.ceil(sum([x.active for x in players])*(1-game.yesProportion[0])+.0001)
    txt = "{} {}'s proposal is available to vote on!\nVote with ~yes or ~no\n{} yes votes will instantly pass the proposal, {} are required to fail it"
    txt = txt.format(playerRole.mention, players[game.turn].name, instPass, instFail)
    await votingChannel.send(txt)
    #Give roles
    await botMember.add_roles(get(nomicServer.roles, name='Game State: Voting'))
    await botMember.remove_roles(get(nomicServer.roles, name='Game State: Proposing'))
    #Begin voting timer
    global voteTask
    game.timerEnd = dt.datetime.now() + dt.timedelta(seconds = game.votingTime)
    voteTask = asyncio.create_task(votingTimeLimit(game.timerEnd))
    await saveData()

@bot.command()
async def transmute(ctx):
    global players, game
    game.transmute = 1
    if not (ctx.channel == votingChannel and game.state == 1 and ctx.author == players[game.turn].discord): return
    game.state = 2
    #FirstVote is whether or not the first vote has been made yet, lastVote is the index of the most recent vote
    game.firstVote = False
    game.lastVote = None
    toVoteRole = get(nomicServer.roles, name='To Vote')
    for player in players:
        if player.active:
            player.currentVote = Vote(0, '', '', player)
            await player.discord.add_roles(toVoteRole)
        else:
            player.currentVote = Vote(-2, '', '', player)
    game.voteNumber = 0
    players[game.turn].stats['proposals'] += 1
    #End timer
    global proposalTask
    proposalTask.cancel()
    if game.rolled:
        global rollTask, rulePool
        rollTask.cancel()
        rule = rulePool[game.rolled-1]
        game.rolled = None
        rule.active = False
        rule.turn = game.globalTurn
        await get(nomicServer.channels, name='rule-pool').fetch_message(rule.msgID).delete()
    #Begin voting phase
    instPass = math.ceil(sum([x.active for x in players])*game.yesProportion[1])
    instFail = math.ceil(sum([x.active for x in players])*(1-game.yesProportion[1])+.0001)
    txt = "{} {}'s proposal is available to vote on!\nVote with ~yes or ~no\n{} yes votes will instantly pass the proposal, {} are required to fail it"
    txt = txt.format(playerRole.mention, players[game.turn].name, instPass, instFail)
    await votingChannel.send(txt)
    #Give roles
    await botMember.add_roles(get(nomicServer.roles, name='Game State: Voting'))
    await botMember.remove_roles(get(nomicServer.roles, name='Game State: Proposing'))
    #Begin voting timer
    global voteTask
    game.timerEnd = dt.datetime.now() + dt.timedelta(seconds = game.votingTime)
    voteTask = asyncio.create_task(votingTimeLimit(game.timerEnd))
    await saveData()

@bot.command()
async def toggleTransmute(ctx):
    global game
    if not (ctx.channel == votingChannel and game.state == 2 and (ctx.author == players[game.turn].discord) or get(nomicServer.roles,name='Historian') in ctx.author.roles): return
    if game.transmute == 0:
        instPass = math.ceil(sum([x.active for x in players])*game.yesProportion[1])
        instFail = math.ceil(sum([x.active for x in players])*(game.yesProportion[1])+.0001)
        txt = 'This proposal does not involve transmutation.\n{} yes votes will instantly pass the proposal, {} are required to fail it'
        await votingChannel.send(txt.format(instPass,instFail))
        game.transmute = 1
    else:
        instPass = math.ceil(sum([x.active for x in players])*game.yesProportion[0])
        instFail = math.ceil(sum([x.active for x in players])*(game.yesProportion[0])+.0001)
        txt = 'This proposal does not involve transmutation.\n{} yes votes will instantly pass the proposal, {} are required to fail it'
        await votingChannel.send(txt.format(instPass,instFail))
        game.transmute = 0



@bot.command()
async def yes(ctx):
    global players, game
    player = get(players, discord = ctx.author)
    if player is None: return
    if not (ctx.channel == votingChannel and game.state == 2 and player.active): return
    if player.currentVote.value == 0:
        player.currentVote = Vote(1, game.voteNumber, dt.datetime.now(), player)
        game.voteNumber += 1
        player = get(players, discord=ctx.author)
        await votingChannel.send("{} has voted!".format(player.name))
        toVoteRole = get(nomicServer.roles, name='To Vote')
        await ctx.author.remove_roles(toVoteRole)
    elif player.currentVote.value != -2:
        await votingChannel.send("You've already voted")
    if players.index(player) != game.turn:
        if not game.firstVote:
            game.firstVote = True
            player.stats['firstVotes'] += 1
        game.lastVote = player
    await checkVotes(0)

@bot.command()
async def no(ctx):
    global players, game
    player = get(players, discord=ctx.author)
    if player is None: return
    if not (ctx.channel == votingChannel and game.state == 2 and player.active): return
    if player.currentVote.value == 0:
        player.currentVote = Vote(2, game.voteNumber, dt.datetime.now(), player)
        game.voteNumber += 1
        await votingChannel.send("{} has voted!".format(player.name))
        toVoteRole = get(nomicServer.roles, name='To Vote')
        await ctx.author.remove_roles(toVoteRole)
    elif player.currentVote.value != -2:
        await votingChannel.send("You've already voted")
    if players.index(player) != game.turn:
        if not game.firstVote:
            game.firstVote = True
            player.stats['firstVotes'] += 1
        game.lastVote = player
    await checkVotes(0)



async def checkVotes(timeUp):
    global game
    allVotes = True
    yesses = 0
    nos = 0
    for player in players:
        if player.currentVote.value == 0 and player.active:
            allVotes = False
        elif player.currentVote.value == 1: yesses += 1
        elif player.currentVote.value == 2: nos += 1
    if not timeUp:
        global summaryMsg
        if summaryMsg: await summaryMsg.delete()
        txt = 'Current votes for/against are {}/{}   ({}%/{}%)\n{}% of all players have voted yes, {}% of all players have voted no'
        txt = txt.format(yesses, nos, round(yesses*100/(yesses+nos),2), round(nos*100/(yesses+nos),2), round(yesses*100/sum([x.active for x in players]),2), round(nos*100/sum([x.active for x in players]),2))
        summaryMsg = await votingChannel.send(txt)

    #All votes have been cast
    if allVotes:
        if yesses/(yesses+nos) >= game.yesProportion[game.transmute] - 0.001:
            await updateChannel.send("All votes have been cast and the proposal has passed. Waiting for the next turn to start")
            await endTurn(1, 0)
        else:
            await updateChannel.send("All votes have been cast and the proposal has failed. Waiting for the next turn to start")
            await endTurn(0, 0)
        return
    #Time is up
    if timeUp:
        if yesses + nos == 0:
            await updateChannel.send("The proposal has failed as no votes were cast, waiting for the next turn to start")
            await endTurn(0, 2)
        elif yesses/(yesses+nos) >= game.timeoutProportion[game.transmute] - 0.001 and (yesses + nos)/sum([x.active for x in players]) >= game.timeoutMinimum[game.transmute] - 0.001:
            await updateChannel.send("Voting time is up, and the proposal has passed. Waiting for the next turn to start")
            await endTurn(1, 2)
        else:
            await updateChannel.send("Voting time is up, and the proposal has failed. Waiting for the next turn to start")
            await endTurn(0, 2)
        return
    #Enough votes to determine a conclusion
    if yesses/sum([x.active for x in players]) >= game.yesProportion[game.transmute] - 0.001:
        await updateChannel.send("There are enough yes votes for the proposal to pass. Waiting for the next turn to start")
        await endTurn(1, 1)
    if nos/sum([x.active for x in players]) > (1-game.yesProportion[game.transmute]) + 0.001:
        await updateChannel.send("There are enough no votes for the proposal to fail. Waiting for the next turn to start")
        await endTurn(0, 1)

async def endTurn(success, endCondition):
    global players, game, summaryMsg
    game.state = 0
    game.timerEnd = None
    yesses = 0
    nos = 0
    for player in players:
        if player.currentVote.value == 1: yesses += 1
        elif player.currentVote.value == 2: nos += 1
    summaryMsg = None
    if endCondition < 3:
        txt = 'Final Votes: {}/{}   ({}%/{}%)   out of {} players'.format(yesses, nos, round(yesses*100/(yesses+nos),2), round(nos*100/(yesses+nos),2), sum([x.active for x in players]))
        await updateChannel.send(txt)
        await votingChannel.send('Voting is now over')

    if endCondition < 2:
        global voteTask
        voteTask.cancel()

    #await endTurnRoles()
    
    if game.lastVote is not None:
        game.lastVote.stats['lastVotes'] += 1

    #Begin waiting phase
    turn = Turn(game.globalTurn)
    turn.proponent = players[game.turn]
    turn.passed = success
    turn.end = endCondition
    turn.voteHistory = []
    for player in players:
        turn.voteHistory.append(player.currentVote)
        player.currentVote = Vote(None, '', '', player)
        #if player.active:
        #    await checkActive(player)
    game.state = 0
    game.firstVote = False
    game.lastVote = None
    game.voteNumber = None
    turns.append(turn)
    
    game.turn += 1
    game.globalTurn += 1
    if game.turn >= len(players):
        game.turn = 0
        await bruhCount()
    while not players[game.turn].active:
        game.turn += 1
        if game.turn > len(players):
            game.turn = 0
            bruhCount()

    if game.ready:
        await start()
    await saveData()

async def endTurnRoles():
    toVoteRole = get(nomicServer.roles, name='To Vote')
    for player in players:
        await player.discord.remove_roles(toVoteRole)
    await players[game.turn].discord.remove_roles(get(nomicServer.roles, name='Current Player'))
    for player in players:
        await player.discord.remove_roles(toVoteRole)
    await botMember.add_roles(get(nomicServer.roles, name='Game State: Waiting'))
    await botMember.remove_roles(get(nomicServer.roles, name='Game State: Proposing'))
    await botMember.remove_roles(get(nomicServer.roles, name='Game State: Voting'))



@bot.event
async def on_message(ctx):
    if 'bruh' in ctx.content.lower():
        await ctx.add_reaction('🇧')
        await ctx.add_reaction('🇷')
        await ctx.add_reaction('🇺')
        await ctx.add_reaction('🇭')
    yesses = ['ja', 'ya', 'yee', 'ye']
    nos = ['nay', 'nah', 'nein', 'mudamudamuda', 'mudamudamudamudamudamudamudamudamudamudamudamudamudamudamudamudamudamudamudamuda', 'wryyy']
    if ctx.content.lower()[1:] in yesses: await yes(ctx)
    if ctx.content.lower()[1:] in nos: await no(ctx)    
    if ctx.author in [x.discord for x in players]:
        player = get(players, discord = ctx.author)
        if player is None: return
        player.lastMessage = dt.datetime.now()
        try:
            player = get(players, discord = ctx.author)
            if ctx.content[0] != '~':
                player.stats['messages'] += 1
                if not player.online:
                    player.online = True
            if ctx.content[:2] == '~#' and int(ctx.content[2:]):
                i = 1
                while ws4.cell(i,1).value is not None:
                    if ws4.cell(i,1).value == int(ctx.content[2:]):
                        await ctx.channel.send(ws4.cell(i,2).value)
                        return
                    i += 1
                await ctx.channel.send('Fool')
                return
        except IndexError:
            pass
        await bot.process_commands(ctx)


async def daily():
    global players
    #Loops once per day
    while True:
        tomorrow = dt.date.today() + dt.timedelta(days=1)
        midnight = dt.datetime.combine(tomorrow, dt.time.min)
        #Loops once per hour
        while True:
            now = dt.datetime.now()
            difference = (midnight-now).total_seconds()
            for player in players:
                if player.active:
                    await checkActive(player)
            if difference < 3600:
                break
            await asyncio.sleep(3600)
            await saveData()
        await asyncio.sleep(difference + 60)
        for player in players:
            if player.active:
                player.stats['daysPlaying'] += 1
            if player.online:
                player.stats['daysOnline'] += 1
                player.online = False
        await saveData()



async def checkActive(player):
    if (dt.datetime.now() - player.lastMessage).total_seconds() < 259200:
        return
    index = players.index(player)
    for turn in turns[-3:]:
        if turn.voteHistory[index].value is not 0:
            return
    await botChannel.send('{} has been made inactive, use ~resurrect to rejoin the game'.format(player.discord.mention))
    await player.discord.remove_roles(playerRole)
    await player.discord.add_roles(get(nomicServer.roles, name='Inactive Player'))
    player.active = False

@bot.command()
async def cryosleep(ctx):
    if not (ctx.author in [x.discord for x in players] and ctx.channel == botChannel):
        return
    player = get(players, discord=ctx.author)
    if player.active:
        player.active = False
        await botChannel.send('You\'ve been made inactive, use ~resurrect to rejoin the game')
        await player.discord.remove_roles(playerRole)
        await player.discord.add_roles(get(nomicServer.roles, name='Inactive Player'))
        if game.state == 2:
            player.currentVote.value = -2
            player.remove_roles(get(nomicServer.roles, name='To Vote'))

@bot.command()
async def resurrect(ctx):
    if not (ctx.author in [x.discord for x in players] and ctx.channel == botChannel):
        return
    player = get(nomicServer.players, discord=ctx.author)
    if not player.active:
        player.active = True
        await botChannel.send('You\'ve rejoined the game!')
        await player.discord.add_roles(playerRole)
        await player.discord.remove_roles(get(nomicServer.roles, name='Inactive Player'))
        if game.state == 2:
            player.currentVote.value = 0
            player.add_roles(get(nomicServer.roles, name='To Vote'))



@bot.command()
async def pool(ctx, text):
    if not(ctx.channel == get(nomicServer.channels, name='rule-crafting') and ctx.author != players[game.turn].discord): return
    global ruleQueue
    total = 0
    for rule in ruleQueue:
        if rule.proposer.discord == ctx.author:
            total += 1
    for rule in rulePool:
        if rule.proposer.discord == ctx.author and rule.active:
            total += 1
    if total >= 2: return
    ruleQueue.append(Rule(len(ruleQueue)+1))
    ruleQueue[-1].text = text
    ruleQueue[-1].proposer = get(players, discord=ctx.author)
    queue = get(nomicServer.channels, name='rule-pool-queue')
    msg = await queue.send('Proposer: ' + ruleQueue[-1].proposer.name + '\n' + ruleQueue[-1].text)
    await msg.add_reaction('\U0001F6D1')
    ruleQueue[-1].msgID = msg.id

@bot.command()
async def roll(ctx):
    global game
    if not(ctx.channel.name=='rule-discussion' and ctx.author == players[game.turn].discord and game.state == 1): return
    number = rnd.random()
    total = sum([x.active] for x in rulePool)
    totalUp = sum([x.up*x.active] for x in rulePool)
    for rule in rulePool:
        if not rule.active: continue
        if (1+rule.up)/(total+totalUp) >= number-0.001:
            game.rolled = rule.number
            await get(nomicServer.channels, name='rule-discussion').send('Selected #{}!\n{}'.format(rule.number, rule.text))
            break
        else:
            number -= (1+rule.up)/(total+totalUp)
    global rollTask
    rollTask = asyncio.create_task(rollTimeLimit(rule))

@bot.command()
async def rule(ctx, number):
    if not(ctx.channel.name == 'rule-discussion' and ctx.author == players[game.turn].discord and game.state == 1): return
    if number[0] != '#': return
    else: number = number[1:]
    rule = get(rulePool, number=number)
    global rollTask
    rollTask = asyncio.create_task(rollTimeLimit(rule))
    

async def rollTimeLimit(rule):
    global game
    game.rolled = rule.number
    await votingChannel.send('{} you have 10 minutes to either propose the rule from the pool or pass')
    await asyncio.sleep(600)
    game.rolled = None
    rule.active = False
    rule.turn = game.globalTurn
    await get(nomicServer.channels, name='rule-pool').fetch_message(rule.msgID).delete()
    await votingChannel.send('~propose\nPool #{} Created by {}\n{}'.format(rule.number, rule.proposer.name, rule.text))

@bot.command()
async def bruh(ctx):
    await ctx.message.add_reaction('🇧')
    await ctx.message.add_reaction('🇷')
    await ctx.message.add_reaction('🇺')
    await ctx.message.add_reaction('🇭')
    await ctx.channel.send('bruh')

async def bruhCount():
    maxlen = max([len(x.name) for x in players])
    txt = '```Bruh Summary:\n'
    for player in players:
        txt +=  '{:{}s}    {}\n'.format(player.name,maxlen+4,player.stats['bruh'])
        player.stats['bruh'] = 0
    txt += '```'
    await updateChannel.send(txt)


@bot.command()
async def transact(ctx, name, number):
    global players, game
    player = get(players, discord=ctx.author)
    player2 = None
    tradingChannel = get(nomicServer.channels,name='trading')
    for x in players:
        if x.name.lower() == name.lower() or x.discord.display_name.lower() == name.lower():
            player2 = x
            break
    if player is None or player2 is None: 
        await tradingChannel.send('Transaction failed to complete')
        return
    number = int(number)
    if not (ctx.channel.name == 'trading' and player.active and player2.active and number <= player.botcoins and number > 0):
        await tradingChannel.send('Transaction failed to complete')
        return
    player.botcoins -= number
    tax = math.ceil(number*0.05)
    player2.botcoins += number - tax
    game.botcoins += tax
    msg = 'Transaction of {} Botcoin from {} to {} successful. {} has been deducted {} Botcoin in Taxation'
    await tradingChannel.send(msg.format(number, player.name, player2.name, player2.name, tax))

@bot.command()
async def account(ctx):
    if not (ctx.channel.name == 'trading' and ctx.author in [x.discord for x in players]): return
    player = get(players, discord=ctx.author)
    await player.discord.send('Your current balance is {}.'.format(player.botcoins))


async def rouletteWait():
    now = dt.datetime.now()
    global minigame, rouletteMsg
    rouletteMsg = None
    await asyncio.sleep((minigame.rouletteTimer - now).total_seconds())
    minigame.rouletteState = 1
    now = dt.datetime.now()
    minigame.rouletteTimer = now + dt.timedelta(days=1)
    await get(nomicServer.channels, name='the-roulette-games').send('Signups are now open!')
    rouletteMsg = await get(nomicServer.channels, name='the-roulette-games').send('```Current Players:\n```')
    await rouletteMsg.pin()
    await rouletteSignup()

async def rouletteSignup():
    now = dt.datetime.now()
    global minigame, rouletteTask
    await asyncio.sleep((minigame.rouletteTimer - now).total_seconds())
    minigame.rouletteState = 2
    if sum([player.stats['roulette']==0 for player in players]) >= 3:
        minigame.rouletteState = 3
        now = dt.datetime.now()
        minigame.rouletteTimer = now + dt.timedelta(days=1)
        await get(nomicServer.channels, name='the-roulette-games').send('Roulette has begun!')
        minigame.rouletteChamber = rnd.randint(1,14)
        rouletteTask = asyncio.create_task(roulettePlay())
    #else:
        #await get(nomicServer.channels, name='the-roulette-games').send('Waiting for 3 players')

async def roulettePlay():
    global minigame, players
    now = dt.datetime.now()
    await asyncio.sleep((minigame.rouletteTimer - now).total_seconds())
    minigame.rouletteState = 0
    minigame.rouletteTimer = now + dt.timedelta(days=1)
    winner = None
    for player in players:
        if player.stats['roulette'] == max([player.stats['roulette'] for player in players]):
            winner = player
            break
    await get(nomicServer.channels, name='the-roulette-games').send('{} has won!'.format(winner.name))
    await rouletteMsg.unpin()
    winner.botcoins += 50 * minigame.rouletteNumber
    for player in players:
        player.stats['roulette'] = -1
    await rouletteWait()

async def joinRoulette(ctx):
    player = get(players,discord=ctx.author)
    if minigame.rouletteState == 0 or minigame.rouletteState == 3 or player is None: return
    if not(player.botcoins >= 50 and player.stats['roulette'] == -1 and player.active): return
    player.botcoins -= 50
    player.stats['roulette'] = 0
    msg = rouletteMsg.content[:-3]
    maxlen = max([len(x.name) for x in players])
    msg += '{:{}s} Alive   {}\n```'.format(player.name + ':', maxlen, player.stats['roulette'])
    await rouletteMsg.edit(content=msg)
    if minigame.rouletteState == 2 and sum([player.stats['roulette']==0 for player in players]) >= 3:
        await get(nomicServer.channels, name='the-roulette-games').send('Roulette has begun!')
        global rouletteTask
        minigame.rouletteChamber = rnd.randint(1,14)
        rouletteTask = asyncio.create_task(roulettePlay())

@bot.command()
async def fire(ctx):
    global minigame, players
    player = get(players,discord=ctx.author)
    if minigame.rouletteState != 3 or player is None or ctx.channel.name != 'the-roulette-games': return
    if player.stats['roulette'] == -1: return
    minigame.rouletteChamber -= 1
    if minigame.rouletteChamber > 0:
        player.stats['roulette'] += 1
        await get(nomicServer.channels, name='the-roulette-games').send('Success!')
        txt = rouletteMsg.content.split('\n')
        msg = ''
        for row in txt:
            if row.startswith(player.name):
                row = row.split('Alive')
                msg += row[0] + 'Alive   ' + str(player.stats['roulette']) + '\n'
            else:
                msg += row + '\n'
        await rouletteMsg.edit(content=msg[:-1])
    else:
        await get(nomicServer.channels, name='the-roulette-games').send('BOOM')
        minigame.rouletteChamber = rnd.randint(1,14)
        player.stats['roulette'] = -1
        txt = rouletteMsg.content.split('\n')
        msg = ''
        for row in txt:
            if row.startswith(player.name):
                row = row.split('Alive')
                msg += row[0] + 'Dead ' + row[1] + '\n'
            else:
                msg += row + '\n'
        await rouletteMsg.edit(content=msg[:-1])
        if sum([player.stats['roulette']>-1 for player in players]) == 1:
            rouletteTask.cancel()
            minigame.rouletteState = 0
            minigame.rouletteTimer = dt.datetime.now() + dt.timedelta(days=1)
            winner = None
            for player in players:
                if player.stats['roulette'] == max([player.stats['roulette'] for player in players]):
                    winner = player
                    break
            await get(nomicServer.channels, name='the-roulette-games').send('{} has won!'.format(winner.name))
            await rouletteMsg.unpin()
            winner.botcoins += 50 * minigame.rouletteNumber
            for player in players:
                player.stats['roulette'] = -1
            await rouletteWait()

@bot.command()
async def joinDebug(ctx, name):
    if ctx.channel.name != 'historian-bot': return
    player = get(players,name=name)
    player.botcoins -= 50
    player.stats['roulette'] = 0
    msg = rouletteMsg.content[:-3]
    msg += '{}: Alive   {}\n```'.format(player.name,player.stats['roulette'])
    await rouletteMsg.edit(content=msg)
    if minigame.rouletteState == 2 and sum([player.stats['roulette']==0 for player in players]) >= 3:
        await get(nomicServer.channels, name='the-roulette-games').send('Roulette has begun!')
        global rouletteTask
        rouletteTask = asyncio.create_task(roulettePlay())

@bot.command()
async def fireDebug(ctx, name):
    if ctx.channel.name != 'historian-bot': return
    global minigame, players
    player = get(players,name=name)
    if minigame.rouletteState != 3 or player is None: return
    if player.stats['roulette'] == -1: return
    minigame.rouletteChamber -= 1
    if minigame.rouletteChamber > 0:
        player.stats['roulette'] += 1
        await get(nomicServer.channels, name='the-roulette-games').send('Success!')
        txt = rouletteMsg.content.split('\n')
        msg = ''
        for row in txt:
            if row.startswith(player.name):
                row = row.split('   ')
                msg += row[0] + '   ' + str(player.stats['roulette']) + '\n'
            else:
                msg += row + '\n'
        await rouletteMsg.edit(content=msg[:-1])
    else:
        await get(nomicServer.channels, name='the-roulette-games').send('BOOM')
        minigame.rouletteChamber = rnd.randint(1,14)
        player.stats['roulette'] = -1
        txt = rouletteMsg.content.split('\n')
        msg = ''
        for row in txt:
            if row.startswith(player.name):
                row = row.split('Alive')
                msg += row[0] + 'Dead' + row[1] + '\n'
            else:
                msg += row + '\n'
        await rouletteMsg.edit(content=msg[:-1])
        if sum([player.stats['roulette']>-1 for player in players]) == 1:
            rouletteTask.cancel()
            minigame.rouletteState = 0
            minigame.rouletteTimer = dt.datetime.now() + dt.timedelta(days=1)
            winner = None
            for player in players:
                if player.stats['roulette'] == max([player.stats['roulette'] for player in players]):
                    winner = player
                    break
            await get(nomicServer.channels, name='the-roulette-games').send('{} has won!'.format(winner.name))
            await rouletteMsg.unpin()
            winner.botcoins += 50 * minigame.rouletteNumber
            for player in players:
                player.stats['roulette'] = -1
            rouletteWait()



bot.loop.create_task(daily())

@bot.event
async def on_command_error(ctx, error):
    if isinstance(error, commands.CommandNotFound):
        return
    raise error

bot.run(token)